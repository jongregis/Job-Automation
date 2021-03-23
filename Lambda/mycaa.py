import json
import boto3
import botocore
import dropbox
import openpyxl as xl
import os
from datetime import datetime
import Levenshtein
import time
from data import cci_programs, commission, au_programs, met_programs, uwlax_programs, csu_programs, tamut_ed4_programs, findMissingClass, set_pricing_cci, set_pricing_au, set_pricing_met, set_pricing_uwlax, set_pricing_csu, set_pricing_tamut_ed4, set_commission, set_pricing_column


def lambda_handler(event, context):
    # month = event['month']
    # current_month = event['current_month']
    month = event['queryStringParameters']['month']
    current_month = event['queryStringParameters']['current_month']
    dbx = dropbox.Dropbox(os.environ['DROPBOX_CODE'])
    s3 = boto3.resource(
        's3', aws_access_key_id=os.environ['AWSAccessKeyId'], aws_secret_access_key=os.environ['AWSSecretKey'])
    pete = "/ECA Back Office/Pete's Backup/MILTARY/PETE ALL 3 SPREADSHEETS MYCAA FOR STACEY AND LISA/MAIN ENROLLMENT FOLDER/SPREADSHEETS/students mycaa FINAL-TODAY.xlsx"

    # Run the Dowload Lambda to get sheets from Dropbox to S3
    # inputForInvoker = {"month": month}
    # client = boto3.client('lambda')
    # response = client.invoke(FunctionName='arn:aws:lambda:us-east-1:077679469516:function:downloadSchoolSpreadsheets', InvocationType='RequestResponse', Payload=json.dumps(inputForInvoker))
    # print('finished invoke')
    # dowload files from S3 Bucket to /tmp dir

    try:
        s3.Bucket('jobautomation').download_file(
            f'{month} 2020.xlsx', f'/tmp/{month} 2020.xlsx')
        s3.Bucket('jobautomation').download_file(
            'students mycaa FINAL-TODAY.xlsx', '/tmp/students mycaa FINAL-TODAY.xlsx')
        s3.Bucket('jobautomation').download_file(
            'Jon weekly email list.xlsx', '/tmp/Jon weekly email list.xlsx')
        print('downloaded to /tmp')
    except botocore.exceptions.ClientError as e:
        if e.response['Error']['Code'] == "404":
            print("The object does not exist.")
        else:
            raise

    monthly_spreadsheet = f"/tmp/{month} 2020.xlsx"
    jon_email_workbook = "/tmp/Jon weekly email list.xlsx"
    pete_spreadsheet = "/tmp/students mycaa FINAL-TODAY.xlsx"

    time_start = datetime.now().strftime("%H:%M:%S")
    print("Starting workbook load at " + time_start)
    wb1 = xl.load_workbook(pete_spreadsheet, read_only=True)
    wb2 = xl.load_workbook(monthly_spreadsheet)
    wb3 = xl.load_workbook(jon_email_workbook)
    time_end = datetime.now().strftime("%H:%M:%S")
    print('Load finished at ' + time_end)

    auburn = wb1["AUBURN & TJC"]
    clemson = wb1["CLEMSON"]
    csu = wb1["COLUMBIA SOUTHERN"]
    lsu = wb1["LOUISIANA STATE"]
    msu = wb1["MONTANA STATE"]
    unh = wb1["NEW HAMPSHIRE"]
    tamu = wb1["TAMUT"]
    wku = wb1["WESTERN KENTUCKY"]
    utep = wb1["UTEP"]
    uwlax = wb1["WISONSIN "]
    desu = wb1["DESU-MyCAA"]
    tamiu = wb1["Texas A&M Interntional"]
    wtamu = wb1["WEST TX A & M"]
    fpu = wb1["FRESNO PACIFIC"]
    monthly = wb2.worksheets[0]
    jon_sheet = wb3.active

    def findNextCell():
        for cell in monthly["C"]:
            if cell.value is None:
                return cell.row
                break
        else:
            return cell.row + 1

    def findNextCellJonEmail():
        for cell in jon_sheet["A"]:

            if cell.value is None:
                return cell.row
                break
        else:
            return cell.row + 1

    def findName(name):
        for cell in monthly["D"]:
            if cell.value == name:
                return True

    def findMissingClass(dictionary, wrong):
        num = 0
        name = ''
        for key in dictionary:
            # print(SequenceMatcher(None, key, 'veterinary assistant specialist').ratio())
            ratio = Levenshtein.ratio(key, wrong)
            if ratio > num:
                num = ratio
                name = key
        if num > 0.5:
            print(
                f'Smart lookup finished. {num}% that {wrong} is {name}')
            return dictionary[name]
        else:
            print("Smart lookup finished. Nothing really seems to match")
            return dictionary[name]

    def nameCleaner(x):
        if '-LAPTOP' in x:
            name, laptop = x.split('-LAPTOP')
            return name
        elif 'LAPTOP' in x:
            name, laptop = x.split('LAPTOP')
            return name
        else:
            return x

    def findDoubleStudent():
        name_list = dict({})
        num = 1
        students = 0
        for i in range(3, 150):
            name = monthly.cell(row=i, column=4).value
            if name == None:
                pass
            else:
                newName = nameCleaner(name)
                if newName in name_list.values():
                    print(f'{newName} is a double name')
                    students += 1
                else:
                    name_list[num] = nameCleaner(name)
                    num += 1
        if students < 1:
            print('no double students found')
            return 'no double students found'
        else:
            return "There are double students"

    def auburn_students(current_month):

        num = findNextCell()
        num1 = findNextCellJonEmail()

        for rowidx, row in enumerate(auburn.rows):

            date = row[2].value
            address = row[6].value
            email = row[7].value
            name = row[8].value
            course = row[9].value
            rep = row[11].value
            vender = row[12].value

            last_number_row = num - 1

            if date and not isinstance(date, str) and date.strftime('%Y') == '2020' and date.strftime('%m') == current_month:
                if findName(name) != True:

                    # place invoice number
                    last_invoice_number = monthly.cell(
                        row=last_number_row, column=11).value
                    monthly.cell(
                        row=num, column=11).value = last_invoice_number+1
                    # place first date
                    date1 = row[0].value
                    date1 = date1.strftime('%m') + '/' + \
                        date1.strftime('%d') + '/' + date1.strftime('%-y')
                    monthly.cell(row=num, column=2).value = date1

                    date = date.strftime('%m') + '/' + \
                        date.strftime('%d') + '/' + date.strftime('%-y')
                    monthly.cell(row=num, column=3).value = date

                    date3 = row[3].value
                    res = isinstance(date3, datetime)

                    if res:
                        date3 = date3.strftime('%m') + '/' + \
                            date3.strftime('%d') + '/' + date3.strftime('%-y')
                        jon_sheet.cell(row=num1, column=3).value = date3
                    else:
                        jon_sheet.cell(row=num1, column=3).value = date3

                    monthly.cell(row=num, column=14).value = address

                    jon_sheet.cell(row=num1, column=2).value = email

                    if 'LAPTOP' in name:
                        monthly.cell(row=num, column=8).value = 'x'
                    monthly.cell(row=num, column=4).value = name
                    jon_sheet.cell(row=num1, column=1).value = name

                    course = course.strip().lower()
                    monthly.cell(row=num, column=5).value = course

                    # code = auburn.cell(row=i, column=11).value
                    # monthly.cell(row=num, column=9).value = code

                    rep = rep.strip().lower()
                    monthly.cell(row=num, column=6).value = rep

                    # setting commision for pete
                    if rep == "maie":
                        monthly.cell(row=num, column=7).value = None
                    elif rep == "pete code lead" and pete_commission() > 5:
                        monthly.cell(row=num, column=7).value = 75
                    elif rep == "pete code lead" and pete_commission() <= 5:
                        monthly.cell(row=num, column=7).value = 'x'
                    else:
                        monthly.cell(
                            row=num, column=7).value = set_commission(course)

                    jon_sheet.cell(row=num1, column=4).value = vender
                    if vender == 'CCI':
                        if course in au_programs:
                            monthly.cell(row=num, column=9).value = 'AU'
                            school = 'AU'
                            monthly.cell(
                                row=num, column=set_pricing_column(
                                    'AU')).value = set_pricing_au(course)
                            price = set_pricing_au(course)
                        elif course not in au_programs:
                            monthly.cell(row=num, column=9).value = 'MET'
                            school = 'MET'
                            monthly.cell(row=num, column=set_pricing_column(
                                'MET')).value = set_pricing_met(course)
                            price = set_pricing_met(course)
                    elif vender == 'Pete Medd':
                        monthly.cell(row=num, column=9).value = 'AU M'
                        school = 'AU M'
                        monthly.cell(row=num, column=set_pricing_column(
                            'MET')).value = set_pricing_met(course)
                        price = set_pricing_met(course)
                    else:
                        monthly.cell(row=num, column=9).value = 'AU ED4'
                        school = 'AU ED4'
                        monthly.cell(row=num, column=set_pricing_column(
                            'AU ED4')).value = set_pricing_met(course)
                        price = set_pricing_met(course)

                    num += 1
                    num1 += 1

        wb2.save(monthly_spreadsheet)


# -----------------------------------------------------------------------------Other Schools-----------------

    def school_tab(current_month, school, schoolString):
        mr = school.max_row
        mc = school.max_column
        num = findNextCell()
        num1 = findNextCellJonEmail()

        for rowidx, row in enumerate(school.rows):
            date = row[2].value
            address = row[6].value
            email = row[7].value
            name = row[8].value
            course = row[9].value
            rep = row[11].value
            vender = row[12].value

            if schoolString == 'TAMIU':
                name = row[7].value

            last_number_row = num - 1
            if date and not isinstance(date, str) and date.strftime('%Y') == '2020' and date.strftime('%m') == current_month:
                if findName(name) != True:
                    # place invoice number
                    last_invoice_number = monthly.cell(
                        row=last_number_row, column=11).value
                    monthly.cell(
                        row=num, column=11).value = last_invoice_number+1
                    # place first date
                    date1 = row[0].value
                    date1 = date1.strftime('%m') + '/' + \
                        date1.strftime('%d') + '/' + date1.strftime('%-y')
                    monthly.cell(row=num, column=2).value = date1

                    date = date.strftime('%m') + '/' + \
                        date.strftime('%d') + '/' + date.strftime('%-y')
                    monthly.cell(row=num, column=3).value = date

                    date3 = row[3].value
                    date3 = date3.strftime('%m') + '/' + \
                        date3.strftime('%d') + '/' + date3.strftime('%-y')
                    jon_sheet.cell(row=num1, column=3).value = date3

                    if schoolString == 'UTEP' or schoolString == 'WTAMU':
                        address = row[7].value

                    elif schoolString == 'TAMIU':
                        address = row[8].value

                    monthly.cell(row=num, column=14).value = address

                    if schoolString == 'UTEP' or schoolString == 'TAMIU' or schoolString == 'WTAMU':
                        email = row[6].value

                    jon_sheet.cell(row=num1, column=2).value = email

                    if 'LAPTOP' in name:
                        monthly.cell(row=num, column=8).value = 'x'
                    monthly.cell(row=num, column=4).value = name
                    jon_sheet.cell(row=num1, column=1).value = name

                    course = course.strip().lower()
                    monthly.cell(row=num, column=5).value = course

                    # checks the rep column for school
                    if schoolString == 'UNH':
                        rep = row[12].value

                    rep = rep.strip().lower()
                    monthly.cell(row=num, column=6).value = rep

                    # setting commision for pete
                    if rep == "maie":
                        monthly.cell(row=num, column=7).value = None
                    elif rep == "pete code lead" and pete_commission() > 5:
                        monthly.cell(row=num, column=7).value = 75
                    elif rep == "pete code lead" and pete_commission() <= 5:
                        monthly.cell(row=num, column=7).value = 'x'
                    else:
                        monthly.cell(
                            row=num, column=7).value = set_commission(course)

                    jon_sheet.cell(row=num1, column=4).value = vender
                    monthly.cell(row=num, column=9).value = schoolString

                    if vender == 'ED4O' and schoolString == 'TAMU' or vender == 'ED40' and schoolString == 'TAMU':
                        monthly.cell(row=num, column=9).value = 'TAMU ED4'
                        monthly.cell(row=num, column=set_pricing_column(
                            'TAMU')).value = set_pricing_tamut_ed4(course)
                        price = set_pricing_tamut_ed4(course)
                    elif vender == 'ED4O' and schoolString == 'DESU':
                        monthly.cell(row=num, column=9).value = 'DESU ED4'
                        monthly.cell(row=num, column=set_pricing_column(
                            'DESU')).value = set_pricing_met(course)
                        price = set_pricing_met(course)
                    elif schoolString == 'CSU':
                        monthly.cell(row=num, column=set_pricing_column(
                            schoolString)).value = set_pricing_csu(course)
                        price = set_pricing_csu(course)
                    elif schoolString == 'UWLAX':
                        monthly.cell(row=num, column=set_pricing_column(
                            schoolString)).value = set_pricing_uwlax(course)
                        price = set_pricing_uwlax(course)
                    elif vender == 'Pete Medd' or vender == 'PETE MEDD':
                        monthly.cell(row=num, column=9).value = 'TAMU M'
                        monthly.cell(row=num, column=set_pricing_column(
                            schoolString)).value = set_pricing_cci(course)
                        price = set_pricing_cci(course)
                    else:
                        monthly.cell(row=num, column=set_pricing_column(
                            schoolString)).value = set_pricing_cci(course)
                        price = set_pricing_cci(course)
                    num += 1
                    num1 += 1

        wb2.save(monthly_spreadsheet)
        print(schoolString)

    def pete_commission():
        num = 0
        for cell in monthly["F"]:
            if cell.value == "pete code lead" or cell.value == 'pete':
                num += 1
        return num

    def runProgram(date):
        try:
            start = findNextCell()
            auburn_students(date)
            school_tab(date, clemson, 'CLEM')
            school_tab(date, csu, 'CSU')
            school_tab(date, lsu, 'LSU')
            school_tab(date, msu, 'MSU')
            school_tab(date, unh, 'UNH')
            school_tab(date, tamu, 'TAMU')
            school_tab(date, wku, 'WKU')
            school_tab(date, uwlax, 'UWLAX')
            school_tab(date, desu, 'DESU')
            school_tab(date, tamiu, 'TAMIU')
            school_tab(date, utep, 'UTEP')
            school_tab(date, wtamu, 'WTAMU')
            school_tab(date, fpu, 'FPU')

            wb2.save(monthly_spreadsheet)
            wb3.save(jon_email_workbook)
            wb1.close()
            end = findNextCell()
            total = end-start
            print("\033[1;32mAll Done Transferring Students!\033[0;0m")
            print("\033[1;32m{} \033[0;0mwere transferred".format(total))
            doubles = findDoubleStudent(month)
            # print(f'Time elapsed: {round(time.time()-start_time,2)} seconds')
            return total, doubles
        except Exception as e:
            print('Something went wrong :-(', e)
            # print(traceback.format_exc())
            return 'Something went wrong', e

    students, double_students = runProgram(current_month)
    # s3.Bucket('jobautomation').upload_file(f'/tmp/{month} 2020.xlsx', f'{month} 2020(updated).xlsx')
    # s3.Bucket('jobautomation').upload_file(jon_email_workbook, 'Jon weekly email list.xlsx')

    overwrite = True
    mode = (dropbox.files.WriteMode.overwrite
            if overwrite
            else dropbox.files.WriteMode.add)

    with open(f'/tmp/{month} 2020.xlsx', 'rb') as f:
        dbx.files_upload(
            f.read(), f'/ECA Back Office/JON/{month} 2020(updated).xlsx', mode, mute=True)
        f.close()
    with open(jon_email_workbook, 'rb') as f:
        dbx.files_upload(
            f.read(), '/ECA Back Office/JON/Jon weekly email list.xlsx', mode, mute=True)
        f.close()

    # response = client.invoke(FunctionName='arn:aws:lambda:us-east-1:077679469516:function:downloadSchoolSpreadsheets', InvocationType='RequestResponse', Payload=json.dumps(inputForInvoker))

    return {
        'statusCode': 200,
        "headers": {'Access-Control-Allow-Origin': '*'},
        'body': json.dumps({'Message': 'Finished Transferring MYCAA Students to Monthly Spreadsheet', 'Transferred_Students': students, 'Number_of_Double_Names': double_students})
    }
