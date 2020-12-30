
import openpyxl as xl
from datetime import datetime


time_start = datetime.now().strftime("%H:%M:%S")
print("Starting workbook load at " + time_start)
jon_email_workbook = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Lisa's Backup/Letters to students/Weekly Email for Lisa/Jon weekly email list.xlsx"
pete_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Pete's Backup/MILTARY/PETE ALL 3 SPREADSHEETS MYCAA FOR STACEY AND LISA/MAIN ENROLLMENT FOLDER/SPREADSHEETS/students mycaa FINAL-TODAY.xlsx"
monthly_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Lisa's Backup/Invoices/2020 Enrollment/Oct 2020.xlsx"
wb1 = xl.load_workbook(pete_spreadsheet)
wb2 = xl.load_workbook(monthly_spreadsheet)
wb3 = xl.load_workbook(jon_email_workbook)


time_end = datetime.now().strftime("%H:%M:%S")
print('Load finished at ' + time_end)
