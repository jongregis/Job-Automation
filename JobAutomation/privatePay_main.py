import openpyxl as xl
from datetime import datetime

monthly_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Lisa's Backup/Invoices/2020 Enrollment/Dec 2020.xlsx"
pete_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Pete's Backup/MILTARY/PETE ALL 3 SPREADSHEETS MYCAA FOR STACEY AND LISA/MAIN ENROLLMENT FOLDER/SPREADSHEETS/students mycaa FINAL-TODAY.xlsx"


wb1 = xl.load_workbook(pete_spreadsheet)
wb2 = xl.load_workbook(monthly_spreadsheet)

DESU = wb1['DESU-CA-PP']
CA = wb1['CREDENTIALLING ASSISTANCE']


monthly = wb2.worksheets[1]


def findNextCell():
    for cell in monthly["C"]:
        if cell.value is None:

            return cell.row
            break
    else:
        return cell.row + 1


def findName(name):
    for cell in monthly["D"]:
        if cell.value == name:
            return True


def school_tab(current_month, school, schoolString, rowNumber):
    mr = school.max_row
    mc = school.max_column
    num = findNextCell()

    for i in range(rowNumber, mr+1):

        c = school.cell(row=i, column=3).value
        if schoolString == 'CA':
            name = school.cell(row=i, column=7).value
        else:
            name = school.cell(row=i, column=8).value
        last_number_row = num - 1

        if c != None and c.strftime('%Y') == '2020' and c.strftime('%m') == current_month:
            if findName(name) != True:
                # place invoice number
                last_invoice_number = monthly.cell(
                    row=last_number_row, column=11).value
                monthly.cell(row=num, column=11).value = last_invoice_number+1
                # place first date
                date1 = school.cell(row=i, column=3).value
                date1 = date1.strftime('%m') + '/' + \
                    date1.strftime('%d') + '/' + date1.strftime('%-y')
                monthly.cell(row=num, column=2).value = date1
                monthly.cell(row=num, column=3).value = date1

                monthly.cell(row=num, column=4).value = name
                course = school.cell(row=i, column=9).value

                course = course.strip().lower()
                monthly.cell(row=num, column=5).value = course

                monthly.cell(row=num, column=9).value = schoolString
                if schoolString == 'CA':

                    misc = school.cell(row=i, column=13).value
                    if misc == '?' or misc == None:
                        misc = '$3000 tuition'
                    else:
                        school_code = school.cell(row=i, column=10).value
                        price, _, *rest = misc.split(' ')
                        _, price = price.split('$')
                        monthly.cell(row=num, column=set_pricing_column(
                            school_code)).value = (int(price) * .75)
                        monthly.cell(
                            row=num, column=9).value = f'CA {school_code}'
                else:
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = cci_programs[course]

                num += 1

    wb2.save(monthly_spreadsheet)


def set_pricing_column(school):

    if school == "DESU":
        return 14
    elif school == "AU":
        return 12
    elif school == "LSUS":
        return 19

    else:
        print("\033[1;31mno school with that name \033[0;0m")


cci_programs = dict({
    "accounting professional": 2016,
    "administrative assistant with quickbooks": 1748,
    "bookkeeping with quickbooks": 1733,
    "business management professional": 1948,
    "childcare specialist": 1942,
    "clinical medical assistant": 1405,
    "clinical medical assistant with ob/gyn": 1405,
    "clinical medical assistant with pediatric specialist": 1405,
    "criminal investigation professional": 2051,
    "dental assisting": 1825,
    "ekg technician cert program": 1250,
    "finanance professional": 1967,
    "finance professional": 1967,
    "human resources professional": 2029,
    "it cyber security professional with comp tia security +": 2050,
    "medical administration assistance": 1250,
    "medical administrative assistant": 1250,
    "medical administrative assistant online": 1250,
    "medical administrative assistant online": 1250,
    "medical billing and coding": 1215,
    "medical billing and coding with medical administrative assistant": 1370,
    "medical billing and coding with medical admin": 1370,
    "medical billing and coding with medical administration": 1370,
    "medical billing & coding w/ medical administrative assistant certificate program includes cmaa and cpc national certification exams": 1370,
    "organizational behavior professional": 2090,
    "paralegal": 1699,
    "paralegal certificate program": 1699,
    "pharmacy technician": 1200,
    "pharmacy technician with medical administration": 1400,
    "pharmacy tech with med admin": 1400,
    "phlebotomy technician": 1575,
    "phlebotomy tech -spanish": 1575,
    "photography entrepreneur with adobe certificate": 1850,
    "photography entrepreneur with adobe": 1850,
    "teachers aide": 2029,
    "veterinary assistant specialist": 1013,
    "welding program": 2050})


def run_program_privatePay(date):
    start = findNextCell()
    school_tab(date, DESU, 'DESU', 10)
    school_tab(date, CA, 'CA', 13)

    wb2.save(monthly_spreadsheet)
    end = findNextCell()
    total = end-start
    print("Done transferring Private Pay Students")
    print("\033[1;32m{} \033[0;0mwere transferred".format(total))
