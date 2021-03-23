import openpyxl as xl
from datetime import datetime
from JobAutomation.data import monthly_spreadsheet

mycaa_invoice = "/Users/jongregis/Python/JobAutomation/JobAutomation/ELearning Automation.xlsm"

wb1 = xl.load_workbook(monthly_spreadsheet)
monthly = wb1.worksheets[2]


wb2 = xl.load_workbook(filename=mycaa_invoice, read_only=False, keep_vba=True)
setup_sheet = wb2.worksheets[3]
data_sheet = wb2.worksheets[1]


def findNextCell():
    for cell in monthly["D"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findNextCellSetup():
    for cell in setup_sheet["A"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findNextCellData():
    for cell in data_sheet["A"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findName(name):
    for cell in setup_sheet["B"]:
        if cell.value == name:
            return True


def findNameData(name):
    for cell in data_sheet["B"]:
        if cell.value == name:
            return True


def move_to_setup_sheet():
    num = findNextCellSetup()

    for i in range(3, findNextCell()):
        name = monthly.cell(row=i, column=4).value
        date = monthly.cell(row=i, column=3).value
        course = monthly.cell(row=i, column=5).value
        school = monthly.cell(row=i, column=9).value
        invoice_number = monthly.cell(row=i, column=11).value
        broward_price = monthly.cell(row=i, column=13).value
        flagler_price = monthly.cell(row=i, column=12).value
        schreiner_price = monthly.cell(row=i, column=14).value
        mns_price = monthly.cell(row=i, column=15).value
        east_ms_price = monthly.cell(row=i, column=16).value
        richmond_price = monthly.cell(row=i, column=17).value
        cleveland_price = monthly.cell(row=i, column=18).value
        green_river_price = monthly.cell(row=i, column=19).value

        if findName(name) != True:

            setup_sheet.cell(row=num, column=1).value = date
            setup_sheet.cell(row=num, column=2).value = name
            setup_sheet.cell(row=num, column=3).value = course
            setup_sheet.cell(row=num, column=4).value = school
            setup_sheet.cell(row=num, column=5).value = invoice_number
            setup_sheet.cell(row=num, column=7).value = name

            if school == "FLAGLER":
                setup_sheet.cell(row=num, column=6).value = flagler_price
            elif school == "BROWARD":
                setup_sheet.cell(row=num, column=6).value = broward_price
            elif school == "SCHREINER":
                setup_sheet.cell(row=num, column=6).value = schreiner_price
            elif school == "MN State":
                setup_sheet.cell(row=num, column=6).value = mns_price
            elif school == "East MS":
                setup_sheet.cell(row=num, column=6).value = east_ms_price
            elif school == "Univ Richmond":
                setup_sheet.cell(row=num, column=6).value = richmond_price
            elif school == "Cleveland":
                setup_sheet.cell(row=num, column=6).value = cleveland_price
            elif school == "Green River":
                setup_sheet.cell(row=num, column=6).value = green_river_price
            else:
                print("School doesnt exist")

            num += 1


def move_to_data_sheet():
    num = findNextCellData()
    mr = setup_sheet.max_row

    for i in range(2, mr+1):
        name = setup_sheet.cell(row=i, column=2).value
        date = setup_sheet.cell(row=i, column=1).value
        course = setup_sheet.cell(row=i, column=3).value
        school = setup_sheet.cell(row=i, column=4).value
        invoice_number = setup_sheet.cell(row=i, column=5).value
        price = setup_sheet.cell(row=i, column=6).value
        real_name = setup_sheet.cell(row=i, column=7).value
        if findNameData(real_name) != True:
            data_sheet.cell(row=num, column=1).value = date
            data_sheet.cell(row=num, column=2).value = real_name
            data_sheet.cell(row=num, column=3).value = course
            data_sheet.cell(row=num, column=4).value = school
            data_sheet.cell(row=num, column=5).value = invoice_number
            data_sheet.cell(row=num, column=6).value = price

            num += 1


def run_docking_invoices_elearning():
    move_to_setup_sheet()
    move_to_data_sheet()
    wb2.save(mycaa_invoice)
    print("\033[1;32mFinished Docking E-Learning Students \033[0;0m")
