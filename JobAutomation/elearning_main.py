import openpyxl as xl
from datetime import datetime
from database.database import execute_query, connection
from JobAutomation.data import monthly_spreadsheet

pete_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Pete's Backup/MILTARY/PETE ALL 3 SPREADSHEETS MYCAA FOR STACEY AND LISA/MAIN ENROLLMENT FOLDER/SPREADSHEETS/ECA ALL SCHOOLS MONTHLY SS.xlsx"


wb1 = xl.load_workbook(pete_spreadsheet)
wb2 = xl.load_workbook(monthly_spreadsheet)

broward = wb1.worksheets[1]
flagler = wb1.worksheets[5]
schreiner = wb1.worksheets[3]
mns = wb1.worksheets[4]
east_ms = wb1['EAST MS CC']
richmond = wb1['UNIV OF RICHMOND']
cleveland = wb1['CLEVELAND STATE UNIV.']
green_river = wb1['GREEN RIVER COLLEGE']

monthly = wb2.worksheets[2]

color_in_hex = schreiner['A32'].fill.start_color
yellow = 'FFFFFF00'


def findNextCell():
    for cell in monthly["C"]:
        if cell.value is None:

            return cell.row
            break
    else:
        return cell.row + 1


def findName(name):
    for cell in monthly["D"]:
        if cell.value == name:
            return True


def school_tab(current_month, school, schoolString, rowNumber, year):
    mr = school.max_row
    mc = school.max_column
    num = findNextCell()

    for i in range(rowNumber, mr+1):

        c = school.cell(row=i, column=3).value
        color_check = school.cell(row=i, column=3)
        name = school.cell(row=i, column=1).value
        last_number_row = num - 1

        if c != None and c != "START DATE" and color_check.fill.start_color.index != yellow and c.strftime('%Y') == year and c.strftime('%m') == current_month:
            if findName(name) != True:

                # place invoice number
                last_invoice_number = monthly.cell(
                    row=last_number_row, column=11).value
                monthly.cell(row=num, column=11).value = last_invoice_number+1
                # place first date
                date1 = school.cell(row=i, column=3).value
                date1 = date1.strftime('%m') + '/' + \
                    date1.strftime('%d') + '/' + date1.strftime('%-y')
                monthly.cell(row=num, column=2).value = date1
                monthly.cell(row=num, column=3).value = date1

                monthly.cell(row=num, column=4).value = name

                if schoolString == "SCHREINER":
                    course = school.cell(row=i, column=8).value
                else:
                    course = school.cell(row=i, column=7).value
                course = course.strip().lower()
                monthly.cell(row=num, column=5).value = course

                monthly.cell(row=num, column=9).value = schoolString
                if schoolString == 'SCHREINER':
                    price = school.cell(row=i, column=10).value
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = price
                else:
                    price = school.cell(row=i, column=9).value
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = price
                first, last = name.split(' ', 1)
                query = f"""
                INSERT INTO Students (first, last, school, course, invoice_number, start_date, amount)
                VALUES ('{first}', '{last}', '{schoolString}', '{course}', '{last_invoice_number+1}', '{date1}', '{price}');
                """
                try:
                    execute_query(connection, query)
                except Exception as e:
                    print(e)

                num += 1

    wb2.save(monthly_spreadsheet)


# print(flagler.cell(row=10, column=3).value)


def set_pricing_column(school):

    if school == "BROWARD":
        return 13
    elif school == "FLAGLER":
        return 12
    elif school == "SCHREINER":
        return 14
    elif school == "MN State":
        return 15
    elif school == "East MS":
        return 16
    elif school == "Univ Richmond":
        return 17
    elif school == "Cleveland":
        return 18
    elif school == "Green River":
        return 19
    else:
        print("\033[1;31mno school with that name \033[0;0m")


def run_program_elearning(date, year):
    start = findNextCell()
    school_tab(date, broward, 'BROWARD', 36, year)
    school_tab(date, flagler, 'FLAGLER', 9, year)
    school_tab(date, schreiner, 'SCHREINER', 22, year)
    school_tab(date, mns, 'MN State', 22, year)
    school_tab(date, east_ms, 'East MS', 14, year)
    school_tab(date, richmond, 'Univ Richmond', 30, year)
    school_tab(date, cleveland, 'Cleveland', 11, year)
    school_tab(date, green_river, 'Green River', 17, year)
    wb2.save(monthly_spreadsheet)
    end = findNextCell()
    total = end-start
    print("Done transferring E-Learning Students")
    print("\033[1;32m{} \033[0;0mwere transferred".format(total))
