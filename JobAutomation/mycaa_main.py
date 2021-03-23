import openpyxl as xl
from datetime import datetime
import traceback
import time
import dropbox
import os
import Levenshtein
from JobAutomation.SortingInvoices.doubleStudent import findDoubleStudent
from JobAutomation.data import cci_programs, commission, au_programs, met_programs, uwlax_programs, csu_programs, tamut_ed4_programs, monthly_spreadsheet
from database.database import execute_query, connection

dbx = dropbox.Dropbox(
    'ucQp1NoOMzUAAAAAAAAAAXYCaTDU29D37vRXCkCwyQ0ep9kcdLbvHFjExMYzesBT')

jon_email_workbook = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Lisa's Backup/Letters to students/Weekly Email for Lisa/Jon weekly email list.xlsx"
pete_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Pete's Backup/MILTARY/PETE ALL 3 SPREADSHEETS MYCAA FOR STACEY AND LISA/MAIN ENROLLMENT FOLDER/SPREADSHEETS/students mycaa FINAL-TODAY.xlsx"

# assert os.path.exists(pete_spreadsheet)
start_time = time.time()
# fileName = 'test1.xlsx'
wb1 = xl.load_workbook(pete_spreadsheet, read_only=True)
auburn = wb1["AUBURN & TJC"]
clemson = wb1["CLEMSON"]
csu = wb1["COLUMBIA SOUTHERN"]
lsu = wb1["LOUISIANA STATE"]
msu = wb1["MONTANA STATE"]
unh = wb1["NEW HAMPSHIRE"]
tamu = wb1["TAMUT"]
wku = wb1["WESTERN KENTUCKY"]
utep = wb1["UTEP"]
uwlax = wb1["WISCONSIN (UWLAX)"]
desu = wb1["DESU-MyCAA"]
tamiu = wb1["Texas A&M Interntional"]
wtamu = wb1["WEST TX A & M"]
fpu = wb1["FRESNO PACIFIC"]

wb2 = xl.load_workbook(monthly_spreadsheet)
monthly = wb2.worksheets[0]

wb3 = xl.load_workbook(jon_email_workbook)
jon_sheet = wb3.active


def findNextCell():
    for cell in monthly["C"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findNextCellJonEmail():
    for cell in jon_sheet["A"]:

        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findName(name):
    for cell in monthly["D"]:
        if cell.value == name:
            return True


def nameCleaner(x):
    if '-LAPTOP' in x:
        name, laptop = x.split('-LAPTOP')
        return name
    elif 'LAPTOP' in x:
        name, laptop = x.split('LAPTOP')
        return name
    else:
        return x


def findMissingClass(dictionary, wrong):
    num = 0
    name = ''
    for key in dictionary:
        # print(SequenceMatcher(None, key, 'veterinary assistant specialist').ratio())
        ratio = Levenshtein.ratio(key, wrong)
        if ratio > num:
            num = ratio
            name = key
    if num > 0.5:
        print(
            f"Smart lookup finished. \033[1;32m{num}%\033[0;0m that \033[1;33m{wrong}\033[0;0m is \033[1;32m{name}\033[0;0m")
        return dictionary[name]
    else:
        print("Smart lookup finished. Nothing really seems to match")
        return dictionary[name]


def auburn_students(current_month, year):

    num = findNextCell()
    num1 = findNextCellJonEmail()

    for rowidx, row in enumerate(auburn.rows):

        date = row[2].value
        address = row[6].value
        email = row[7].value
        name = row[8].value
        laptop = row[9].value
        course = row[10].value
        rep = row[12].value
        vender = row[13].value

        last_number_row = num - 1

        if date and not isinstance(date, str) and date.strftime('%Y') == year and date.strftime('%m') == current_month:
            if findName(name) != True:

                # place invoice number
                last_invoice_number = monthly.cell(
                    row=last_number_row, column=11).value
                monthly.cell(row=num, column=11).value = last_invoice_number+1
                # place first date
                date1 = row[0].value
                date1 = date1.strftime('%m') + '/' + \
                    date1.strftime('%d') + '/' + date1.strftime('%-y')
                monthly.cell(row=num, column=2).value = date1

                date = date.strftime('%m') + '/' + \
                    date.strftime('%d') + '/' + date.strftime('%-y')
                monthly.cell(row=num, column=3).value = date

                date3 = row[3].value
                res = isinstance(date3, datetime)

                if res:
                    date3 = date3.strftime('%m') + '/' + \
                        date3.strftime('%d') + '/' + date3.strftime('%-y')
                    jon_sheet.cell(row=num1, column=3).value = date3
                else:
                    jon_sheet.cell(row=num1, column=3).value = date3

                monthly.cell(row=num, column=14).value = address

                jon_sheet.cell(row=num1, column=2).value = email

                if laptop == 'Y':
                    monthly.cell(row=num, column=8).value = 'x'
                monthly.cell(row=num, column=4).value = name
                jon_sheet.cell(row=num1, column=1).value = name

                course = course.strip().lower()
                monthly.cell(row=num, column=5).value = course

                # code = auburn.cell(row=i, column=11).value
                # monthly.cell(row=num, column=9).value = code

                rep = rep.strip().lower()
                monthly.cell(row=num, column=6).value = rep

                # setting commision for pete
                if rep == "maie":
                    monthly.cell(row=num, column=7).value = None
                elif rep == "pete code lead" and pete_commission() > 5:
                    monthly.cell(row=num, column=7).value = 75
                elif rep == "pete code lead" and pete_commission() <= 5:
                    monthly.cell(row=num, column=7).value = 'x'
                else:
                    monthly.cell(
                        row=num, column=7).value = set_commission(course)

                jon_sheet.cell(row=num1, column=4).value = vender
                if vender == 'CCI':
                    if course in au_programs:
                        monthly.cell(row=num, column=9).value = 'AU'
                        school = 'AU'
                        monthly.cell(
                            row=num, column=set_pricing_column(
                                'AU')).value = set_pricing_au(course)
                        price = set_pricing_au(course)
                    elif course not in au_programs:
                        monthly.cell(row=num, column=9).value = 'MET'
                        school = 'MET'
                        monthly.cell(row=num, column=set_pricing_column(
                            'MET')).value = set_pricing_met(course)
                        price = set_pricing_met(course)
                elif vender == 'Pete Medd':
                    monthly.cell(row=num, column=9).value = 'AU M'
                    school = 'AU M'
                    monthly.cell(row=num, column=set_pricing_column(
                        'MET')).value = set_pricing_met(course)
                    price = set_pricing_met(course)
                else:
                    monthly.cell(row=num, column=9).value = 'AU ED4'
                    school = 'AU ED4'
                    monthly.cell(row=num, column=set_pricing_column(
                        'AU ED4')).value = set_pricing_met(course)
                    price = set_pricing_met(course)
                # name = nameCleaner(name)
                # first, last = name.split(' ', 1)
                # query = f"""
                # INSERT INTO Students (first, last, school, course, email, address, rep, invoice_number, start_date, amount)
                # VALUES ('{first}', '{last}', '{school}', '{course}', '{email}', '{address}', '{rep}', '{last_invoice_number+1}', '{date}', '{price}');
                # """
                # try:
                #     execute_query(connection, query)
                # except Exception as e:
                #     print(e)

                num += 1
                num1 += 1

    wb2.save(monthly_spreadsheet)
    wb3.save(jon_email_workbook)


# -----------------------------------------------------------------------------Other Schools-----------------

def school_tab(current_month, school, schoolString, year):
    mr = school.max_row
    mc = school.max_column
    num = findNextCell()
    num1 = findNextCellJonEmail()

    for rowidx, row in enumerate(school.rows):
        date = row[2].value
        address = row[6].value
        email = row[7].value
        name = row[8].value
        laptop = row[9].value
        course = row[10].value
        rep = row[12].value
        vender = row[13].value

        if schoolString == 'TAMIU':
            name = row[7].value

        last_number_row = num - 1
        if date and not isinstance(date, str) and date.strftime('%Y') == year and date.strftime('%m') == current_month:
            if findName(name) != True:
                # place invoice number
                last_invoice_number = monthly.cell(
                    row=last_number_row, column=11).value
                monthly.cell(row=num, column=11).value = last_invoice_number+1
                # place first date
                date1 = row[0].value
                date1 = date1.strftime('%m') + '/' + \
                    date1.strftime('%d') + '/' + date1.strftime('%-y')
                monthly.cell(row=num, column=2).value = date1

                date = date.strftime('%m') + '/' + \
                    date.strftime('%d') + '/' + date.strftime('%-y')
                monthly.cell(row=num, column=3).value = date

                date3 = row[3].value
                date3 = date3.strftime('%m') + '/' + \
                    date3.strftime('%d') + '/' + date3.strftime('%-y')
                jon_sheet.cell(row=num1, column=3).value = date3

                if schoolString == 'UTEP' or schoolString == 'WTAMU':
                    address = row[7].value

                elif schoolString == 'TAMIU':
                    address = row[10].value

                monthly.cell(row=num, column=14).value = address

                if schoolString == 'UTEP' or schoolString == 'TAMIU' or schoolString == 'WTAMU':
                    email = row[6].value
                if schoolString == 'TAMIU':
                    email = row[7].value

                jon_sheet.cell(row=num1, column=2).value = email

                if laptop == 'Y':
                    monthly.cell(row=num, column=8).value = 'x'
                monthly.cell(row=num, column=4).value = name
                jon_sheet.cell(row=num1, column=1).value = name

                course = course.strip().lower()
                monthly.cell(row=num, column=5).value = course

                # checks the rep column for school
                if schoolString == 'UNH':
                    rep = row[12].value

                rep = rep.strip().lower()
                monthly.cell(row=num, column=6).value = rep

                # setting commision for pete
                if rep == "maie":
                    monthly.cell(row=num, column=7).value = None
                elif rep == "pete code lead" and pete_commission() > 5:
                    monthly.cell(row=num, column=7).value = 75
                elif rep == "pete code lead" and pete_commission() <= 5:
                    monthly.cell(row=num, column=7).value = 'x'
                else:
                    monthly.cell(
                        row=num, column=7).value = set_commission(course)

                jon_sheet.cell(row=num1, column=4).value = vender
                monthly.cell(row=num, column=9).value = schoolString

                if vender == 'ED4O' and schoolString == 'TAMU' or vender == 'ED40' and schoolString == 'TAMU':
                    monthly.cell(row=num, column=9).value = 'TAMU ED4'
                    monthly.cell(row=num, column=set_pricing_column(
                        'TAMU')).value = set_pricing_tamut_ed4(course)
                    price = set_pricing_tamut_ed4(course)
                elif vender == 'ED4O' and schoolString == 'DESU':
                    monthly.cell(row=num, column=9).value = 'DESU ED4'
                    monthly.cell(row=num, column=set_pricing_column(
                        'DESU')).value = set_pricing_met(course)
                    price = set_pricing_met(course)
                elif schoolString == 'CSU':
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = set_pricing_csu(course)
                    price = set_pricing_csu(course)
                elif schoolString == 'UWLAX':
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = set_pricing_uwlax(course)
                    price = set_pricing_uwlax(course)
                elif vender == 'Pete Medd' or vender == 'PETE MEDD':
                    monthly.cell(row=num, column=9).value = 'TAMU M'
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = set_pricing_cci(course)
                    price = set_pricing_cci(course)
                else:
                    monthly.cell(row=num, column=set_pricing_column(
                        schoolString)).value = set_pricing_cci(course)
                    price = set_pricing_cci(course)
                # try:
                #     name = nameCleaner(name)
                #     first, last = name.split(' ', 1)
                #     query = f"""
                #     INSERT INTO Students (first, last, school, course, email, address, rep, invoice_number, start_date, amount)
                #     VALUES ('{first}', '{last}', '{schoolString}', '{course}', '{email}', '{address}', '{rep}', '{last_invoice_number+1}', '{date}', '{price}');
                #     """

                #     execute_query(connection, query)
                # except Exception as e:
                #     print(f'Problem inserting to database: {e}')

                num += 1
                num1 += 1

    wb2.save(monthly_spreadsheet)
    wb3.save(jon_email_workbook)
    print(schoolString)


def set_pricing_cci(course):
    if course in cci_programs:

        return cci_programs[course]
        print(cci_programs[course])
    else:
        print("\033[1;31mNo class found for CCI, smart update started... \033[0;0m")
        return findMissingClass(cci_programs, course)


def set_pricing_au(course):
    if course in au_programs:

        return au_programs[course]
    else:
        print(
            "\033[1;31mNo class update pricing for au, smart update started... \033[0;0m")
        return findMissingClass(au_programs, course)


def set_pricing_met(course):
    if course in met_programs:
        return met_programs[course]
    else:
        print(
            "\033[1;31mNo class update pricing for met, smart update started... \033[0;0m")
        return findMissingClass(met_programs, course)


def set_pricing_uwlax(course):
    if course in uwlax_programs:
        return uwlax_programs[course]
    else:
        print(
            "\033[1;31mNo class update pricing for uwlax, smart update started... \033[0;0m")
        return findMissingClass(uwlax_programs, course)


def set_pricing_csu(course):
    if course in csu_programs:
        return csu_programs[course]
    else:
        print(
            "\033[1;31mNo class, update pricing for csu, smart update started... \033[0;0m")
        return findMissingClass(csu_programs, course)


def set_pricing_tamut_ed4(course):
    if course in tamut_ed4_programs:
        return tamut_ed4_programs[course]
    else:
        print(
            "\033[1;31mNo class, update pricing for TAMUT ED4, smart update started... \033[0;0m")
        return findMissingClass(tamut_ed4_programs, course)


def set_pricing_column(school):

    if school == "AU ED4":
        return 12
    elif school == "AU":
        return 13
    elif school == "MET":
        return 12
    elif school == "WKU":
        return 17
    elif school == "LSU":
        return 19
    elif school == "UTEP":
        return 20
    elif school == "CLEM":
        return 21
    elif school == "UWLAX":
        return 22
    elif school == "CSU":
        return 25
    elif school == "TAMU":
        return 26
    elif school == "MSU":
        return 28
    elif school == "UNH":
        return 31
    elif school == "DESU":
        return 32
    elif school == "TAMIU":
        return 33
    elif school == "WTAMU":
        return 30
    elif school == "FPU":
        return 24
    else:
        print("\033[1;31mno school with that name \033[0;0m")


def pete_commission():
    num = 0
    for cell in monthly["F"]:
        if cell.value == "pete code lead" or cell.value == 'pete':
            num += 1
    return num


def set_commission(course):
    if course in commission:
        return commission[course]
    else:
        print('')


def runProgram(date, month, year):
    try:
        start = findNextCell()
        auburn_students(date, year)
        school_tab(date, clemson, 'CLEM', year)
        school_tab(date, csu, 'CSU', year)
        school_tab(date, lsu, 'LSU', year)
        school_tab(date, msu, 'MSU', year)
        school_tab(date, unh, 'UNH', year)
        school_tab(date, tamu, 'TAMU', year)
        school_tab(date, wku, 'WKU', year)
        school_tab(date, uwlax, 'UWLAX', year)
        school_tab(date, desu, 'DESU', year)
        school_tab(date, tamiu, 'TAMIU', year)
        school_tab(date, utep, 'UTEP', year)
        school_tab(date, wtamu, 'WTAMU', year)
        school_tab(date, fpu, 'FPU', year)

        wb2.save(monthly_spreadsheet)
        wb3.save(jon_email_workbook)
        wb1.close()
        end = findNextCell()
        total = end-start
        print("\033[1;32mAll Done Transferring Students!\033[0;0m")
        print("\033[1;32m{} \033[0;0mwere transferred".format(total))
        doubles = findDoubleStudent(month)
        print(f'Time elapsed: {round(time.time()-start_time,2)} seconds')
        return total, doubles
    except Exception as e:
        print('Something went wrong :-(', e)
        print(traceback.format_exc())
        return 'Something went wrong', e


# runProgram('10')

# findNextCellPete(auburn)
# set_pricing_cci('Veteriary Assistant')
