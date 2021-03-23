import openpyxl as xl
from datetime import datetime
import os
from JobAutomation.data import monthly_spreadsheet

mycaa_invoice = "/Users/jongregis/Python/JobAutomation/JobAutomation/MYCAA Automation.xlsm"

wb1 = xl.load_workbook(monthly_spreadsheet)
monthly = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=mycaa_invoice, read_only=False, keep_vba=True)
setup_sheet = wb2.worksheets[3]
data_sheet = wb2.worksheets[1]


def findNextCell():
    for cell in monthly["D"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findNextCellSetup():
    for cell in setup_sheet["A"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findNextCellData():
    for cell in data_sheet["A"]:
        if cell.value is None:
            return cell.row
            break
    else:
        return cell.row + 1


def findName(name):
    for cell in setup_sheet["B"]:
        if cell.value == name:
            return True


def findNameData(name):
    for cell in data_sheet["B"]:
        if cell.value == name:
            return True


def move_to_setup_sheet():
    num = findNextCellSetup()

    for i in range(3, findNextCell()):
        name = monthly.cell(row=i, column=4).value
        date = monthly.cell(row=i, column=3).value
        course = monthly.cell(row=i, column=5).value
        school = monthly.cell(row=i, column=9).value
        invoice_number = monthly.cell(row=i, column=11).value
        au_price = monthly.cell(row=i, column=13).value
        met_price = monthly.cell(row=i, column=12).value
        ed4_price = monthly.cell(row=i, column=12).value
        wku_price = monthly.cell(row=i, column=17).value
        lsu_price = monthly.cell(row=i, column=19).value
        clem_price = monthly.cell(row=i, column=21).value
        uwlax_price = monthly.cell(row=i, column=22).value
        csu_price = monthly.cell(row=i, column=25).value
        tamu_price = monthly.cell(row=i, column=26).value
        msu_price = monthly.cell(row=i, column=28).value
        unh_price = monthly.cell(row=i, column=31).value
        desu_price = monthly.cell(row=i, column=32).value
        utep_price = monthly.cell(row=i, column=20).value
        tamiu_price = monthly.cell(row=i, column=33).value
        wtamu_price = monthly.cell(row=i, column=30).value
        fpu_price = monthly.cell(row=i, column=24).value

        if findName(name) != True:

            setup_sheet.cell(row=num, column=1).value = date
            setup_sheet.cell(row=num, column=2).value = name
            setup_sheet.cell(row=num, column=3).value = course
            setup_sheet.cell(row=num, column=4).value = school
            setup_sheet.cell(row=num, column=5).value = invoice_number
            setup_sheet.cell(row=num, column=7).value = nameCleaner(name)

            if school == "AU":
                setup_sheet.cell(row=num, column=6).value = au_price
            elif school == "MET":
                setup_sheet.cell(row=num, column=6).value = met_price
            elif school == "AU M":
                setup_sheet.cell(row=num, column=6).value = met_price
            elif school == "TAMU M":
                setup_sheet.cell(row=num, column=6).value = tamu_price
            elif school == "AU ED4":
                setup_sheet.cell(row=num, column=6).value = ed4_price
            elif school == "WKU":
                setup_sheet.cell(row=num, column=6).value = wku_price
            elif school == "LSU":
                setup_sheet.cell(row=num, column=6).value = lsu_price
            elif school == "CLEM":
                setup_sheet.cell(row=num, column=6).value = clem_price
            elif school == "UWLAX":
                setup_sheet.cell(row=num, column=6).value = uwlax_price
            elif school == "CSU":
                setup_sheet.cell(row=num, column=6).value = csu_price
            elif school == "TAMU":
                setup_sheet.cell(row=num, column=6).value = tamu_price
            elif school == "MSU":
                setup_sheet.cell(row=num, column=6).value = msu_price
            elif school == "UNH":
                setup_sheet.cell(row=num, column=6).value = unh_price
            elif school == "DESU":
                setup_sheet.cell(row=num, column=6).value = desu_price
            elif school == "TAMU ED4":
                setup_sheet.cell(row=num, column=6).value = tamu_price
            elif school == "DESU ED4":
                setup_sheet.cell(row=num, column=6).value = desu_price
            elif school == "UTEP":
                setup_sheet.cell(row=num, column=6).value = utep_price
            elif school == "TAMIU":
                setup_sheet.cell(row=num, column=6).value = tamiu_price
            elif school == "WTAMU":
                setup_sheet.cell(row=num, column=6).value = wtamu_price
            elif school == "FPU":
                setup_sheet.cell(row=num, column=6).value = fpu_price
            else:
                print("\033[1;31mSchool doesnt exist!\033[0;0m")

            num += 1


def nameCleaner(x):
    if '-LAPTOP' in x:
        name, laptop = x.split('-LAPTOP')
        return name
    elif 'LAPTOP' in x:
        name, laptop = x.split('LAPTOP')
        return name
    else:
        return x


def move_to_data_sheet():
    num = findNextCellData()
    mr = setup_sheet.max_row

    for i in range(2, mr+1):
        name = setup_sheet.cell(row=i, column=2).value
        date = setup_sheet.cell(row=i, column=1).value
        course = setup_sheet.cell(row=i, column=3).value
        school = setup_sheet.cell(row=i, column=4).value
        invoice_number = setup_sheet.cell(row=i, column=5).value
        price = setup_sheet.cell(row=i, column=6).value
        real_name = setup_sheet.cell(row=i, column=7).value
        if findNameData(real_name) != True:
            data_sheet.cell(row=num, column=1).value = date
            data_sheet.cell(row=num, column=2).value = real_name
            data_sheet.cell(row=num, column=3).value = course
            data_sheet.cell(row=num, column=4).value = school
            data_sheet.cell(row=num, column=5).value = invoice_number
            data_sheet.cell(row=num, column=6).value = price

            num += 1


def quickEditName():
    mr = setup_sheet.max_row
    for i in range(2, mr+1):
        name = setup_sheet.cell(row=i, column=2).value
        setup_sheet.cell(row=i, column=7).value = nameCleaner(name)
    wb2.save(mycaa_invoice)


def run_docking_invoices():
    move_to_setup_sheet()
    move_to_data_sheet()
    wb2.save(mycaa_invoice)
    print("\033[1;32mFinished docking MYCAA students \033[0;0m")
