import openpyxl as xl
from datetime import datetime
import os

massage_invoice_sheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Pete's Backup/MILTARY/PETE ALL 3 SPREADSHEETS MYCAA FOR STACEY AND LISA/MAIN ENROLLMENT FOLDER/SPREADSHEETS/MASSAGE MONTHLY SS for Jon.xlsx"

wb1 = xl.load_workbook(massage_invoice_sheet)

payout_tab = wb1.active
mr = payout_tab.max_row

color_in_hex = payout_tab['A172'].fill.start_color
green = 'FF92D050'

total = 0
for i in range(170, 199):
    name = payout_tab.cell(row=i, column=1).value
    payment1 = payout_tab.cell(row=i, column=12)
    payment2 = payout_tab.cell(row=i, column=13)
    payment3 = payout_tab.cell(row=i, column=14)
    payment4 = payout_tab.cell(row=i, column=15)
    payment5 = payout_tab.cell(row=i, column=16)
    payment6 = payout_tab.cell(row=i, column=17)
    payment7 = payout_tab.cell(row=i, column=18)
    payment8 = payout_tab.cell(row=i, column=19)

    tuition = payout_tab.cell(row=i, column=6).value
    payment_type = payout_tab.cell(row=i, column=2).value
    if name == None:
        pass
    elif 'INVOIICING' in name:
        pass
    elif 'Students' not in name:
        if payment1.fill.start_color.index == green and '?' not in payment1.value and payment2.fill.start_color.index != green:
            amount, payment = payment1.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment2.fill.start_color.index == green and payment3.fill.start_color.index != green:
            amount, payment = payment2.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment3.fill.start_color.index == green and payment4.fill.start_color.index != green:
            amount, payment = payment3.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment4.fill.start_color.index == green and payment5.fill.start_color.index != green:
            amount, payment = payment4.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment5.fill.start_color.index == green and payment6.fill.start_color.index != green:
            amount, payment = payment5.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment6.fill.start_color.index == green and payment7.fill.start_color.index != green:
            amount, payment = payment6.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment7.fill.start_color.index == green and payment8.fill.start_color.index != green:
            amount, payment = payment7.value.split('(')
            if '$' in amount:
                sign, amount = amount.split('$')
                total += float(amount)
            else:
                total += float(amount)
        elif payment_type == "PRIVATE PAY-PIF(Paid in Full)":
            tuition = int(tuition)
            tuition = (tuition * 0.15) + 1250
            total += tuition
