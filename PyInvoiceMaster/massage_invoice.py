from datetime import datetime, date
from models import InvoiceInfo, ServiceProviderInfo, ClientInfo, Item, Transaction
from templateMassage import SimpleInvoiceMassage
from dateutil.relativedelta import *
# from .pp_Invoice import create_invoice_PrivatePay
import openpyxl as xl


massage_invoice_sheet = "/Users/jongregis/Desktop/MASSAGE MONTHLY SS for Jon-USE THIS.xlsx"
monthly_spreadsheet = "/Volumes/SanDisk Extreme SSD/Dropbox (ECA Consulting)/ECA Back Office/Lisa's Backup/Invoices/2020 Enrollment/Oct 2020.xlsx"

wb1 = xl.load_workbook(massage_invoice_sheet)
wb2 = xl.load_workbook(monthly_spreadsheet)

payout_tab = wb1.active
monthly = wb2.worksheets[1]
mr = payout_tab.max_row

color_in_hex = payout_tab['A323'].fill.start_color
green = 'FF92D050'
gold = 'FFFFC000'
blank = '00000000'

total = 0


def findNextCell():
    for cell in monthly["C"]:
        if cell.value is None:

            return cell.row
            break
    else:
        return cell.row + 1


def findName(name):
    for cell in monthly["D"]:
        if cell.value == name:
            return True


def create_invoice_Massage(month, name, description, tuition, percentage, school, invoice_number):

    doc = SimpleInvoiceMassage(
        f'/Users/jongregis/Desktop/test/Massage Invoice {invoice_number} {month}.pdf')

    # Paid stamp, optional
    doc.is_paid = True

    current_date = datetime.now().strftime('%m'+'/'+'%d'+'/'+'%-y')
    due_date = datetime.now() + relativedelta(days=+15)
    due_date = due_date.strftime('%m'+'/'+'%d'+'/'+'%-y')

    doc.invoice_info = InvoiceInfo(
        invoice_number, current_date, due_date)

    # Service Provider Info, optional
    doc.service_provider_info = ServiceProviderInfo(
        name='771 Cool Creek Rd LLC',
        street='1257 Water St',
        city='Wrightsville',
        state='PA',
        # country='My Country',
        post_code='17368',

        # vat_tax_number='Vat/Tax number'
    )

    # Client info, optional
    doc.client_info = ClientInfo(name=name, school=school)

    # Calculate Tuition

    # Add Item

    for i in range(362, 399):
        name = payout_tab.cell(row=i, column=1).value
        name_color = payout_tab.cell(row=i, column=1)
        payment1 = payout_tab.cell(row=i, column=12)
        payment2 = payout_tab.cell(row=i, column=13)
        payment3 = payout_tab.cell(row=i, column=14)
        payment4 = payout_tab.cell(row=i, column=15)
        payment5 = payout_tab.cell(row=i, column=16)
        payment6 = payout_tab.cell(row=i, column=17)
        payment7 = payout_tab.cell(row=i, column=18)
        payment8 = payout_tab.cell(row=i, column=19)

        tuition = payout_tab.cell(row=i, column=6).value
        payment_type = payout_tab.cell(row=i, column=2).value
        school_col = payout_tab.cell(row=i, column=3).value

        if name_color.fill.start_color.index == blank or name == None:
            continue
        elif 'INVOIICING' in name:
            continue
        elif 'MYCAA' not in payment_type and 'SALLIE MAE' not in payment_type:
            print(name)
            start_date = payout_tab.cell(row=i, column=4).value
            start_date = start_date.strftime('%m') + '/' + \
                start_date.strftime('%d') + '/' + start_date.strftime('%-y')
            if payment_type == "PRIVATE PAY-PIF(Paid in Full)":
                tuition = int(tuition)
                tuition = (tuition * 0.15) + 1250
                doc.add_item(Item(start_date, name, tuition,
                                  'PRIVATE PAY-PIF'))
            elif payment_type == "TAMUT-PIF":
                tuition = int(tuition)
                tuition = (tuition * 0.15) + 1250
                doc.add_item(Item(start_date, name, tuition,
                                  'TAMUT-PIF'))
            elif payment_type == "AHCI-PIF" or payment_type == 'AHCI-MD Apprenticeship-PIF':
                tuition = 800
                doc.add_item(Item(start_date, name, tuition,
                                  'AHCI-PIF'))
            elif payment_type == "PAID IN FULL-PAID PETE MEDD":
                tuition = (int(tuition) * 0.15)+450
                doc.add_item(Item(start_date, name, tuition,
                                  'PIF GILLIAN'))
            elif payment1.fill.start_color.index == green and '?' not in payment1.value and payment2.fill.start_color.index != green:
                amount, payment = payment1.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(
                        Item(start_date, name, amount, f'(1 of {check_payments(i)})'))
                else:
                    doc.add_item(
                        Item(start_date, name, amount, f'(1 of {check_payments(i)})'))
            elif payment2.fill.start_color.index == green and payment3.fill.start_color.index != green:
                amount, payment = payment2.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(2 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(2 of {check_payments(i)})'))
            elif payment3.fill.start_color.index == green and payment4.fill.start_color.index != green:
                amount, payment = payment3.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(3 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(3 of {check_payments(i)})'))
            elif payment4.fill.start_color.index == green and payment5.fill.start_color.index != green:
                amount, payment = payment4.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(4 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(4 of {check_payments(i)})'))
            elif payment5.fill.start_color.index == green and payment6.fill.start_color.index != green:
                amount, payment = payment5.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(5 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(5 of {check_payments(i)})'))
            elif payment6.fill.start_color.index == green and payment7.fill.start_color.index != green:
                amount, payment = payment6.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(6 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(6 of {check_payments(i)})'))
            elif payment7.fill.start_color.index == green and payment8.fill.start_color.index != green:
                amount, payment = payment7.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(7 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(7 of {check_payments(i)})'))
            elif payment5.fill.start_color.index == gold and payment6.fill.start_color.index != gold:
                amount, payment = payment5.value.split('(')
                if '$' in amount:
                    sign, amount = amount.split('$')
                    doc.add_item(Item(start_date, name, amount,
                                      f'(5 of {check_payments(i)})'))
                else:
                    doc.add_item(Item(start_date, name, amount,
                                      f'(5 of {check_payments(i)})'))
        elif name_color.fill.start_color.index == green and 'SALLIE MAE' in payment_type:
            if findName(name) != True:
                num = findNextCell()
                last_number_row = num - 1
                last_invoice_number = monthly.cell(
                    row=last_number_row, column=11).value
                monthly.cell(row=num,
                             column=11).value = last_invoice_number+1
                create_invoice_PrivatePay(start_date, name, 'PIF Massage',
                                          tuition*.75, '25%', f'{school_col} M', last_invoice_number+1, 'PP')
                monthly.cell(row=num, column=4).value = name
                monthly.cell(row=num, column=3).value = start_date
                monthly.cell(row=num, column=5).value = 'PIF Massage'
                if school_col == 'Auburn':
                    monthly.cell(row=num, column=9).value = 'AU M'
                elif school_col == 'TAMUT':
                    monthly.cell(row=num, column=9).value = 'TAMUT M'
                monthly.cell(row=num, column=11).value = last_invoice_number+1
                if school_col == 'Auburn':
                    monthly.cell(row=num, column=15).value = tuition*.75
                elif school_col == 'TAMUT':
                    monthly.cell(row=num, column=16).value = tuition*.75
                wb2.save(monthly_spreadsheet)
        else:
            continue

    # Optional
    doc.set_bottom_tip(
        "Email: paul.fears@psfinternational.com<br /><strong>Make All Checks Payable To 771 Cool Creek Rd, LLC</strong><br/>Thank You For Your Bussiness!")

    doc.finish()


def check_payments(row):
    payment = payout_tab.cell(row=row, column=12).value
    first, second = payment.split('of ')
    num, part = second.split(')')
    return num


create_invoice_Massage('11.1.20', '',
                       '', '', '', '', '027')
# print(color_in_hex)
