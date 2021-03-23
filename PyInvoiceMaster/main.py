from .sampleInvoice import create_invoice, create_invoice_Ed4, create_invoice_PSF
from .e_learning_invoice import create_invoice_ELearning
from .pp_Invoice import create_invoice_PrivatePay
import openpyxl as xl


mycaa_invoice = "/Users/jongregis/Python/JobAutomation/JobAutomation/MYCAA Automation.xlsm"
elearning_invoices = "/Users/jongregis/Python/JobAutomation/JobAutomation/ELearning Automation.xlsm"
privatePay_invoices = "/Users/jongregis/Python/JobAutomation/JobAutomation/PP Automation.xlsm"
wb2 = xl.load_workbook(filename=mycaa_invoice, read_only=False, keep_vba=True)
wb3 = xl.load_workbook(filename=elearning_invoices,
                       read_only=False, keep_vba=True)
wb4 = xl.load_workbook(filename=privatePay_invoices,
                       read_only=False, keep_vba=True)

data_sheet = wb2.worksheets[1]
elearning_sheet = wb3.worksheets[1]
privatePay_sheet = wb4.worksheets[1]


# MYCAA main run for invoice creation


def excel_to_pdf():
    mr = data_sheet.max_row
    for i in range(2, mr+1):
        start_date = data_sheet.cell(row=i, column=1).value
        name = data_sheet.cell(row=i, column=2).value
        description = data_sheet.cell(row=i, column=3).value
        school = data_sheet.cell(row=i, column=4).value
        invoice_number = data_sheet.cell(row=i, column=5).value
        total = data_sheet.cell(row=i, column=6).value
        if school == 'DESU ED4':
            create_invoice_Ed4(start_date, name, description,
                               total, '', school, invoice_number)
        elif school == 'TAMU M':
            create_invoice_PSF(start_date, name, description,
                               total, '', school, invoice_number)
        elif school == 'TAMU ED4':
            school = 'TAMUT'
            create_invoice_PSF(start_date, name, description,
                               total, '', school, invoice_number)
        else:
            create_invoice(start_date, name, description,
                           total, '', school, invoice_number)

    print("MYCAA PDF Inovices Done!")

# E-Learning main run for invoice creation


def excel_to_pdf_ELearning():
    mr = elearning_sheet.max_row

    for i in range(2, mr+1):
        start_date = elearning_sheet.cell(row=i, column=1).value
        name = elearning_sheet.cell(row=i, column=2).value
        description = elearning_sheet.cell(row=i, column=3).value
        school = elearning_sheet.cell(row=i, column=4).value
        invoice_number = elearning_sheet.cell(row=i, column=5).value
        total = elearning_sheet.cell(row=i, column=6).value
        create_invoice_ELearning(start_date, name, description,
                                 total, '', school, invoice_number)

    print("E-Learning PDF Inovices Done!")


def excel_to_pdf_PrivatePay():
    mr = privatePay_sheet.max_row

    for i in range(2, mr+1):
        start_date = privatePay_sheet.cell(row=i, column=1).value
        name = privatePay_sheet.cell(row=i, column=2).value
        description = privatePay_sheet.cell(row=i, column=3).value
        school = privatePay_sheet.cell(row=i, column=4).value

        invoice_number = privatePay_sheet.cell(row=i, column=5).value
        total = privatePay_sheet.cell(row=i, column=6).value
        create_invoice_PrivatePay(start_date, name, description,
                                  total, '', school, invoice_number)
        print(name)

    print("Private Pay PDF Inovices Done!")
